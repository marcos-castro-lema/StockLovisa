import traceback
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import subprocess
import sys

# Global variable for region
selected_region = "AU 10"
region_buffer = 0
print("Initial selected region:", selected_region)
region_suffix = selected_region.split()[0] if selected_region else ""
region_buffer =  int(selected_region.split()[1]) if selected_region else 0
print("Selected region suffix:", region_suffix)
print("Selected region buffer:", region_buffer)

region_buffer_map = {"MY": 10, "US": 20, "AU": 30}

def load_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        excel_label.config(text=os.path.basename(file_path))
        app.excel_path = file_path

def load_json_file():
    file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    if file_path:
        json_label.config(text=os.path.basename(file_path))
        app.json_path = file_path

def update_progress(step, total_steps):
    progress_bar['value'] = (step / total_steps) * 100
    app.update_idletasks()

def apply_sheet_formatting(ws):
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws['A2']

def apply_conditional_formatting(ws):
    qty_diff_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == 'QTY_Diff':
            qty_diff_col_idx = idx
            break
    if qty_diff_col_idx:
        col_letter = get_column_letter(qty_diff_col_idx)
        rule = ColorScaleRule(start_type='num', start_value=0, start_color='00FF00',
                              end_type='num', end_value=100, end_color='FF0000')
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{ws.max_row}', rule)

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

def compare_and_generate_report():
    global selected_region
    try:
        if not selected_region:
            messagebox.showerror("Error", "Please select a region before generating the report.")
            return

        total_steps = 6
        step = 0
        update_progress(step, total_steps)

        # Load Excel (Shopify)
        df_excel = pd.read_excel(app.excel_path, engine='openpyxl')
        required_excel_cols = {'Variant SKU', 'Variant Inventory Qty', 'Status'}
        if not required_excel_cols.issubset(df_excel.columns):
            raise ValueError("Excel file is missing required columns.")

        # Robust search for Commited in Warehouse column
        commited_candidates = [col for col in df_excel.columns if "committed" in col.lower() and "warehouse" in col.lower()]
        commited_col = commited_candidates[0] if commited_candidates else None
        if not commited_col:
            raise ValueError("Excel file does not contain a column with both 'Commited' and 'Warehouse' in its name.")

        df_excel = df_excel[['Variant SKU', 'Variant Inventory Qty', 'Status', commited_col]].rename(columns={
            'Variant SKU': 'SKU',
            'Variant Inventory Qty': 'QTY',
            commited_col: 'Commited in Warehouse'
        })
        step += 1
        update_progress(step, total_steps)

        # Load JSON (WMS)
        with open(app.json_path, 'r') as f:
            raw_json = json.load(f)
        inventory_list = raw_json["INVENTORYSNAPSHOT"]["warehouse"]["INVENTORY"]
        df_json = pd.DataFrame(inventory_list)
        required_json_cols = {'SKU', 'QTYAVAILABLE'}
        if not required_json_cols.issubset(df_json.columns):
            raise ValueError("JSON file is missing required fields.")
        df_json = df_json[['SKU', 'QTYAVAILABLE']].rename(columns={'QTYAVAILABLE': 'QTY'})
        step += 1
        update_progress(step, total_steps)

        # Convert SKU to string
        df_excel['SKU'] = df_excel['SKU'].astype(str)
        df_json['SKU'] = df_json['SKU'].astype(str)

        # Merge and calculate difference
        merged_df = pd.merge(df_json, df_excel, on='SKU', how='outer', suffixes=('_WMS', '_Shopify'), indicator=True)
        merged_df['QTY_Diff'] = (merged_df['QTY_WMS'] - merged_df['QTY_Shopify']).abs()
        merged_df['QTY_Match'] = merged_df['QTY_WMS'] == merged_df['QTY_Shopify']
        merged_df['Status'] = merged_df['Status'].fillna('N/A')
        merged_df.sort_values(by='QTY_Diff', ascending=True, inplace=True)
        step += 1
        update_progress(step, total_steps)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = f"WMS_Shopify_{selected_region}_Stock_Crosscheck_{timestamp}.xlsx"

        # Prepare sheets
        both_df = merged_df[merged_df['_merge'] == 'both'].copy()
        warehouse_df = merged_df[merged_df['_merge'] == 'left_only'].copy()
        shopify_df = merged_df[merged_df['_merge'] == 'right_only'].copy()

        # BOTH tab logic
        both_df['Commited in Warehouse'] = both_df['Commited in Warehouse']
        both_df['FINAL_QTY_Diff'] = both_df['QTY_Diff'] - both_df['Commited in Warehouse']
        both_df['ACTION/COMMENTS'] = ""

        selected_region_prefix = selected_region.split()[0] if selected_region else ""
        selected_region_buffer = region_buffer_map.get(selected_region_prefix, 0)
    
        def comment_logic(row):
            if row['QTY_WMS'] <= selected_region_buffer and row['QTY_Shopify'] == 0:
                return f"Its OK, its country buffer defined {selected_region}"
            elif row['FINAL_QTY_Diff'] > 0:
                return "Check the differences between systems"
            return ""
        both_df['ACTION/COMMENTS'] = both_df.apply(comment_logic, axis=1)

        # WAREHOUSE ONLY
        warehouse_df['ACTION/COMMENTS'] = "Article in WMS > check with Product team article should be online or not"

        # SHOPIFY ONLY logic
        def shopify_comment(row):
            if row['Status'] == 'Active' and row['QTY_Shopify'] != 0:
                return "Set to zero in OMS due to WMS transferout not notified"
            return ""
        shopify_df['ACTION/COMMENTS'] = shopify_df.apply(shopify_comment, axis=1)

        # Save to Excel
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            suffix = f"_{selected_region}"
            merged_df.to_excel(writer, sheet_name=f'Full_Comparison{suffix}', index=False)
            both_df.to_excel(writer, sheet_name=f'BOTH{suffix}', index=False)
            warehouse_df.to_excel(writer, sheet_name=f'WAREHOUSE ONLY{suffix}', index=False)
            shopify_df.to_excel(writer, sheet_name=f'SHOPIFY ONLY{suffix}', index=False)
        step += 1
        update_progress(step, total_steps)

        # Apply formatting and auto-adjust column widths
        wb = load_workbook(report_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            apply_sheet_formatting(ws)
            apply_conditional_formatting(ws)
            auto_adjust_column_width(ws)
        wb.save(report_path)
        step += 1
        update_progress(step, total_steps)

        # Open file
        if sys.platform == "win32":
            os.startfile(report_path)
        elif sys.platform == "darwin":
            subprocess.call(["open", report_path])
        else:
            subprocess.call(["xdg-open", report_path])

        progress_bar['value'] = 0
    except Exception as e:
        with open("error_log.txt", "a") as log:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log.write(f"--- ERROR on {timestamp} ---\n")
            log.write(f"Error Type: {type(e).__name__}\n")
            log.write(f"Error Message: {e}\n")
            log.write(traceback.format_exc()) # <-- Esto añade el stack trace completo
            log.write("---------------------------------\n\n")
        progress_bar['value'] = 0
        messagebox.showerror("Error", str(e))

# GUI setup
app = tk.Tk()
app.title("Stock Comparison Tool")
app.geometry("400x420")
app.configure(bg="white")

style = ttk.Style(app)
style.theme_use('clam')
style.configure("TButton", font=('Segoe UI', 10), padding=6)
style.configure("TLabel", font=('Segoe UI', 10), background="white", foreground="black")
style.configure("TProgressbar", thickness=20)

# Dropdown for region selection
ttk.Label(app, text="Select Region:").pack(pady=10)
region_var = tk.StringVar()
region_dropdown = ttk.Combobox(app, textvariable=region_var, values=["AU 10","US 20","MY 0"], state="readonly")
region_dropdown.pack()
region_dropdown.current(0)

selected_label = ttk.Label(app, text="Current Region: AU")
selected_label.pack(pady=5)

def set_region(*args):
    global selected_region
    selected_region = region_var.get()
    selected_label.config(text=f"Current Region: {selected_region}")

region_var.trace_add("write", set_region)
selected_region = region_var.get()

# Existing widgets
ttk.Label(app, text="Load Shopify Excel File:").pack(pady=10)
excel_btn = ttk.Button(app, text="Browse Excel", command=load_excel_file)
excel_btn.pack()
excel_label = ttk.Label(app, text="No file selected")
excel_label.pack()

ttk.Label(app, text="Load WMS JSON File:").pack(pady=5)
json_btn = ttk.Button(app, text="Browse JSON", command=load_json_file)
json_btn.pack()
json_label = ttk.Label(app, text="No file selected")
json_label.pack()

compare_btn = ttk.Button(app, text="Compare and Generate Report", command=compare_and_generate_report)
compare_btn.pack(pady=10)

progress_bar = ttk.Progressbar(app, mode='determinate', maximum=100)
progress_bar.pack(pady=10, fill='x', padx=20)

app.excel_path = None
app.json_path = None

print("Stock Comparison Tool is running...")
app.mainloop()